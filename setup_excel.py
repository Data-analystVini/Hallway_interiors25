from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection

# Headers
headers = [
    "Entry Type", "Date", "Project Name", "Amount By", "Amount",
    "Invoice Number", "Party Name", "Expense For", "Description", "Labour Name"
]

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Entries"

# Add headers
ws.append(headers)

# Freeze top row
ws.freeze_panes = 'A2'

# Style headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="5A3E2B", end_color="5A3E2B", fill_type="solid")
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill

# Set column widths (optional but helpful)
widths = [15, 12, 20, 15, 12, 18, 18, 18, 25, 20]
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65 + i)].width = w

# Protect sheet
ws.protection.sheet = True
ws.protection.password = "1234"

# Save the file
wb.save("data.xlsx")
print("✅ Excel setup completed successfully!")
